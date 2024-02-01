import re
import yaml
import os
import glob
import csv
from benedict import benedict
from loguru import logger
from slugify import slugify
from openpyxl import load_workbook


# global cache
_global_cache = {}


def additional(device_defaults, device_facts, ciscoconf, onboarding_config):
    """add device properties to the device properties

    Add the beginning our response is empty. At the end this dict contains the additional values that are added to nautobot
    """
    basedir = onboarding_config.get('git').get('app_configs').get('path')
    directory = os.path.join(basedir, './onboarding/additional_values/')
    files = []

    # init response
    response = benedict(keyattr_dynamic=True)

    logger.debug(f'reading config from {directory} for adding additional values')
    # we read all *.yaml files in our additional_values data config dir
    for filename in glob.glob(os.path.join(directory, "*.yaml")):
        logger.debug(f'reading additional config {filename.rsplit("/")[-1]}')

        config = read_file(filename, device_defaults.get('platform'))
        if config is None:
            continue

        # add filename to our list of files that were processed
        files.append(os.path.basename(filename))

        for item_config in config.get('additional'):
            name = item_config.get('name','unknown')
            logger.debug(f'got additional config; name={name}')
            if 'file' in item_config:
                # it is either a csv or an xlsx file
                get_additional_values_from_file(
                    response, 
                    item_config,
                    device_facts,
                    device_defaults,
                    onboarding_config)
            else:
                get_additional_values_from_config(
                    response,
                    device_facts, 
                    device_defaults, 
                    item_config, 
                    ciscoconf)
    return response

def get_additional_values_from_file(response, item_config, device_facts, device_defaults, onboarding_config):
    """return additional values values read from a file"""
    file_format = item_config.get('format','csv')

    if file_format == 'csv':
        return add_values_from_csv(response, item_config, device_facts, device_defaults, onboarding_config)
    elif file_format == 'excel' or file_format == 'xlsx':
        return add_values_from_excel(response, item_config, device_facts, device_defaults, onboarding_config)
    else:
        logger.error(f'unknown file format {file_format}')
        return response

def add_values_from_csv(response, item_config, device_facts, device_defaults, onboarding_config):

    basedir = onboarding_config.get('git').get('app_configs').get('path')
    directory = os.path.join(basedir, './onboarding/additional_values/')
    
    filename = "%s/%s" % (directory, item_config.get('file'))
    # check if file exists
    if not os.path.isfile(filename):
        logger.error(f'file {filename} does not exists. Please correct your config.')
        return

    logger.debug(f'reading additional CSV values from {filename}')

    # set default values
    delimiter = item_config.get('delimiter',',')
    quotechar = item_config.get('quotechar','|')
    quoting_cf = item_config.get('quoting','minimal')
    newline = item_config.get('newline','')
    if quoting_cf == "none":
        quoting = csv.QUOTE_NONE
    elif quoting_cf == "all":
        quoting = csv.QUOTE_ALL
    elif quoting_cf == "nonnumeric":
        quoting = csv.QUOTE_NONNUMERIC
    else:
        quoting = csv.QUOTE_MINIMAL
    logger.info(f'reading mapping {filename} delimiter={delimiter} ' \
                 'quotechar={quotechar} newline={newline} quoting={quoting_cf}')

    # read CSV file
    with open(filename, newline=newline) as csvfile:
        csvreader = csv.DictReader(csvfile, delimiter=delimiter, quoting=quoting, quotechar=quotechar)
        for row in csvreader:
            # maybe there are multiple items
            for matches_on in item_config.get('matches_on'):
                # sot key => name of key in our sot
                # csv_key => name of column in our csv file
                for sot_key, csv_key in matches_on.items():
                    df = device_facts.get(sot_key)
                    if df and df == row.get(csv_key):
                        del row[csv_key]
                        for k,v in row.items():
                            response[k] = v
                    dd = device_defaults.get(sot_key)
                    if dd and dd == row.get(csv_key):
                        del row[csv_key]
                        for k,v in row.items():
                            response[k] = v
    csvfile.close()
    return response

def add_values_from_excel(response, item_config, device_facts, device_defaults, onboarding_config):

    # read excel file and add values to our response if a certain value matches
    # eg. col 1 of our excel sheet is named 'hostname'. We are now checking 
    # if the device_facts or the device_defaults have a key called hostname
    # If the key was found and the value matches (device_facts hostname == excel value) 
    # we add all the values of the sheet to our response
    #
    # the item_config contains a mapping sot_key => excel_key
    # eg. device_facts = {'xxx': 'myhost'}
    # than the mapping [{'xxx':'hostname'}] maps the key hostname of our
    # excel sheet to the key xxx of our device_facts or device_defaults

    global _global_cache

    table = []

    basedir = onboarding_config.get('git').get('app_configs').get('path')
    directory = os.path.join(basedir, './onboarding/additional_values/')

    filename = "%s/%s" % (directory, item_config.get('file'))
    # check if file exists
    if not os.path.isfile(filename):
        logger.error(f'file {filename} does not exists. Please correct your config.')
        return

    matching_key = item_config.get('matches_on')
    logger.debug(f'reading additional XLSX values from {filename} matching_key: {matching_key}')

    if filename in _global_cache:
        workbook = _global_cache.get(filename)
    else:
        # Load the workbook
        workbook = load_workbook(filename = filename)
        _global_cache[filename] = workbook

    # Select the active worksheet
    worksheet = workbook.active
    
    # loop through table and build list of dict
    rows = worksheet.max_row
    columns = worksheet.max_column + 1 
    for row in range(2, rows + 1):
        line = {}
        for col in range(1, columns):
            key = worksheet.cell(row=1, column=col).value
            value = worksheet.cell(row=row, column=col).value
            line[key] = value
        table.append(line)
    
    for row in table:
        # maybe there are multiple items
        for matches_on in matching_key:
            # sot key => name of key in our sot
            # excel_key => name of column in our csv file
            for sot_key, excel_key in matches_on.items():
                df = device_facts.get(sot_key,'')
                if len(df) > 0 and df.lower() == row.get(excel_key,'').lower():
                    logger.debug(f'sot_key: {sot_key} excel_key: {excel_key} found in device_facts')
                    # remove value (and only the value that matches)
                    # from our row / otherwise the key will be added to our response dict 
                    del row[excel_key]
                    # add all values to our response dict
                    for key,value in row.items():
                        # do not add None or empty values
                        if value and len(value) > 0:
                            response[key] = value
                dd = device_defaults.get(sot_key,'')
                if len(dd) > 0 and dd.lower() == row.get(excel_key,'').lower():
                    logger.debug(f'sot_key: {sot_key} excel_key: {excel_key} found in device_defaults')
                    del row[excel_key]
                    for key,value in row.items():
                        # do not add None or empty values
                        if isinstance(value, str):
                            if value and len(value) > 0:
                                response[key] = value
                        elif value:
                            response[key] = value

    logger.debug('processed XLSX file successfully')

def get_additional_values_from_config(response, device_facts, device_defaults, item_config, ciscoconf):
    """Checks whether the device meets the configured criteria.
       If the device meets the criteria the additional values are added to the device 
       property.
    """
    matches = get_matches(
        device_facts, 
        device_defaults, 
        item_config.get('matches',{}), 
        ciscoconf)
    if not matches:
        return
    # build dict using key and values configured in 'values'
    for key, value in item_config.get('values').items():
        # logger.debug(f'key {key} value {value}')
        if isinstance(value, str):
            if '__named__' in value:
                group = value.split('__named__')[1]
                response[key] = matches.groups(group)[0]
            else:
                response[key] = value
        elif isinstance(value, dict):
            response[key] = {}
            for k, v in value.items():
                #logger.debug(f'k {k} v {v}')
                if '__named__' in v:
                    groups = v.split('__named__')
                    string = ""
                    for g in groups:
                        if len(g) == 0:
                            continue
                        if g in matches.groupdict():
                            string += matches.group(g)
                        else:
                            string += g
                    response[key][k] = string
                else:
                    response[key][k] = v
                if k == "slug":
                    response[key][k] = slugify(response[key][k])
        else:
            response[key] = value

def get_matches(device_facts, device_defaults, matches, ciscoconf):
    """
    loop through ALL matches and check if it matches
    
    get_matches looks either at the (global/interface) config, the facts,
    or the default values of the device. lookups like ic (case insenitive) 
    or re (use regular expression) can be used. 

    examples:

    facts__fqdn__re: k(?P<digits>\d+)rt
    facts__hostname__ic: myhostname
    config__global__ic: username my_user
    config__interfaces__ic: ip address

    get_matches returns the value that matches
    """
    logger.debug('looping through all matches in config file')
    for name, value in matches.items():
        logger.debug(f'analyzing name={name}')
        if '__' in name:
            splits = name.split('__')
            source = key = lookup = ""
            if len(splits) == 3:
                # source / key / lookup
                source = splits[0]
                key = splits[1]
                lookup = splits[2]
            elif len(splits) == 2:
                source = splits[0]
                key = splits[1]
            # logger.debug(f'source: {source} key: {key} lookup: {lookup}')

            if source == "facts":
                obj = device_facts.get(key)
            elif source == "defaults":
                obj = device_defaults.get(key)
            elif source == "config":
                # look if value is found in config
                if lookup != "":
                    match = "match__%s" % lookup
                else:
                    match = "match"
                props = {match: value, 'ignore_leading_spaces': True}
                if key == "global" and ciscoconf:
                    return ciscoconf.find_in_global(props)
                elif key == "interfaces" and ciscoconf:
                    return ciscoconf.find_in_interfaces(props)
                else:
                    logger.error('unknown key; must be global or interfaces')
                    continue
            else:
                logger.error(f'no source found or source {source} invalid')
                continue

            if lookup == '':
                if obj == value:
                    # logger.debug(f'exact match on {key}')
                    return obj
            elif 'ci' == lookup or 'ic' == lookup:
                # logger.debug(f'ci lookup found on {key}')
                if value.lower() in obj.lower():
                    return obj
            elif 're' == lookup or 'rei' == lookup:
                # logger.debug(f'regular expression {value} on {key} found')
                if obj is not None:
                    if 'rei' == lookup:
                        p = re.compile(value, re.IGNORECASE)
                        m = p.search(obj)
                    else:
                        p = re.compile(value)
                        m = p.search(obj)
                    if m:
                        #logger.debug(f'regular expression matches on {obj}')
                        return m
    return False

def read_file(filename, device_platform):
    """read yaml file and check if file must be processed (is active and platform matches)"""
    with open(filename) as f:
        config = {}
        logger.debug(f'open file {filename.rsplit("/")[-1]}')
        try:
            config = yaml.safe_load(f.read())
            if config is None:
                logger.error("could not parse file %s" % filename)
                return None
        except Exception as exc:
            logger.error("could not read file %s; got exception %s" % (filename, exc))
            return None
        name = config.get('name')
        platform = config.get('platform')

        if not config.get('active'):
            logger.debug(f'file {filename.rsplit("/")[-1]} is not active')
            return None
        if platform is not None:
            if platform != 'all' and platform != device_platform:
                logger.debug("skipping custom field %s wrong platform %s" % (name, platform))
                return None
        logger.debug('config read and parsed successfully')
        return config
