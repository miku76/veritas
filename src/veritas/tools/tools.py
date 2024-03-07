import yaml
import os
import getpass
import pytricia
import hashlib
import smtplib
import datetime
import re
from loguru import logger
from openpyxl import load_workbook

# veritas
import veritas.auth


def get_miniapp_config(appname:str, app_path:str, config_file:str=None, subdir:str="miniapps") -> dict | None:
    """return config of miniapp

    Priority:
        1. prio: user specified file (if absolute path)
        2. prio: config in home directory ~/.veritas
        3. prio: config in local directory
        4. prio: config in local ./conf/ directory
        5. prio: config in /etc/veritas/

    Parameters
    ----------
    appname : str
        name of app
    app_path : str
        path to app config
    config_file : str, optional
        name of config, by default None
    subdir : str, optional
        name of subdir, by default "miniapps"

    Returns
    -------
    config
        YAML object containing config
    """    
    
    config_filename = config_file if config_file else f'{appname}.yaml'

    abs_path_config = config_file if config_file and config_file.startswith('/') else None
    homedir_config_file = f'{os.path.expanduser("~")}/.veritas/{subdir}/{appname}/{config_filename}'
    local_config_file = f'{app_path}/{config_filename}'
    local_subdir_config_file = f'{app_path}/conf/{config_filename}'
    etc_config_file = f'/etc/veritas/{subdir}/{appname}/{config_filename}'

    if abs_path_config and os.path.exists(abs_path_config):
        filename = abs_path_config
    elif os.path.exists(homedir_config_file):
        filename = homedir_config_file
    elif os.path.exists(local_config_file):
        filename = local_config_file
    elif os.path.exists(local_subdir_config_file):
        filename = local_subdir_config_file
    elif os.path.exists(etc_config_file):
        filename = etc_config_file
    else:
        logger.critical(f'neither {abs_path_config} nor {homedir_config_file}, {local_config_file} or {local_subdir_config_file} ' \
                        f'or {etc_config_file} exist')
        return None

    try:
        with open(filename) as f:
            return yaml.safe_load(f.read())
    except Exception as exc:
        logger.error(f'could not read or parse config; got exception {exc}')
        return None

def get_value_from_dict(dictionary:dict, keys:list) -> tuple[list|str|dict|int] | None:
    """get value from dict

    Parameters
    ----------
    dictionary : dict
        the source dict
    keys : list
        list of keys to follow

    Returns
    -------
    value
        the value (can be list, dict, str, int ....)
    """
    if dictionary is None:
        return None

    nested_dict = dictionary

    for key in keys:
        try:
            nested_dict = nested_dict[key]
        except KeyError:
            return None
        except IndexError:
            return None
        except TypeError:
            return nested_dict

    return nested_dict

def get_value_from_dict_and_list(dictionary:dict, keys:list) -> tuple[list|str|dict|int] | None:
    """get value from dict and list (the list can be a value of dict)

    Parameters
    ----------
    dictionary : dict
        the source dictionary
    keys : list
        list of keys to follow

    Returns
    -------
    value
        the value (can be list, dict, str, int ....)
    """
    if dictionary is None:
        return None

    nested_dict = dictionary

    for key in keys:
        try:
            nested_dict = nested_dict[key]
        except KeyError:
            return None
        except IndexError:
            return None
        except TypeError:
            # check if it is a list
            if isinstance(nested_dict, list):
                values = []
                remaining_keys = []
                found = False
                for i in keys:
                    if i == key:
                        found = True
                    if found:
                        remaining_keys.append(i)
                for i in nested_dict:
                    d = get_value_from_dict(i, remaining_keys)
                    values.append(d)
            return values

    return nested_dict

def remove_key_from_dict(dictionary:dict, key:str, key_in_str:bool=False) -> dict:
    """remove key from dict

    Parameters
    ----------
    dictionary : dict
        the source dict
    key : str
        name of key to remove
    key_in_str : bool, optional
        if True, the key is in a string, by default False

    Returns
    -------
    value
        the value (can be list, dict, str, int ....)
    """
    if dictionary is None:
        return None

    if isinstance(dictionary, list):
        for i in dictionary:
            remove_key_from_dict(i, key, key_in_str)
    else:
        for k, v in dict(dictionary).items():
            if isinstance(v, dict):
                remove_key_from_dict(v, key, key_in_str)
            elif isinstance(v, list):
                for i in v:
                    remove_key_from_dict(i, key, key_in_str)
            else:
                if key_in_str and key in k:
                    del dictionary[k]
                else:
                    if k == key:
                        del dictionary[k]

def convert_arguments_to_properties(*unnamed, **named) -> dict | str | list:
    """convert named and unnamed arguments to a single dict

    Returns
    -------
    properties : dict
        the 'dict' that contains all values
    """
    properties = {}
    if len(unnamed) > 0:
        for param in unnamed:
            if isinstance(param, dict):
                for key,value in param.items():
                    properties[key] = value
            elif isinstance(param, str):
                # it is just a text like log('something to log')
                return param
            elif isinstance(param, tuple):
                for tup in param:
                    if isinstance(tup, dict):
                        for key,value in tup.items():
                            properties[key] = value
                    elif isinstance(tup, str):
                        return tup
                    elif isinstance(tup, list):
                        return tup
            elif isinstance(param, list):
                return param
            elif param is None:
                pass
            else:
                logger.error(f'cannot use paramater {param} / {type(param)} as value')
    for key,value in named.items():
            properties[key] = value
    
    return properties

def get_username_and_password(profile_config:dict, profile_name:str=None, 
                              cfg_username:str=None, cfg_password:str=None) -> tuple[str, str]:
    """return username and password

    Parameters
    ----------
    config : dict
        config
    profile : str, optional
        name of profile, by default None
    cfg_username : str, optional
        configured username (mostly args.username), by default None
    cfg_password : str, optional
        configured password (mostly args.password), by default None

    Returns
    -------
    username, password : list
        the username and the password
    """    
    username = None
    password = None

    if profile_name is not None:
        username = profile_config.get('profiles',{}).get(profile_name,{}).get('username')
        token = profile_config.get('profiles',{}).get(profile_name,{}).get('password')
        if username and token:
            password = veritas.auth.decrypt(token=token,
                                            encryption_key=os.getenv('ENCRYPTIONKEY'), 
                                            salt=os.getenv('SALT'), 
                                            iterations=int(os.getenv('ITERATIONS')))

    # overwrite username and password if configured by user
    username = cfg_username if cfg_username else username
    password = cfg_password if cfg_password else password

    username = input("Username (%s): " % getpass.getuser()) if not username else username
    password = getpass.getpass(prompt="Enter password for %s: " % username) if not password else password

    return username, password

def read_excel_file(filename:str) -> list:
    """read excel file and return content as list

    Parameters
    ----------
    filename : str
        filename of the excel file

    Returns
    -------
    table : list
        list of rows
    """
    table = []

    # Load the workbook
    workbook = load_workbook(filename = filename)

    # Select the active worksheet
    worksheet = workbook.active
    
    # loop through table and build list of dict
    rows = worksheet.max_row
    columns = worksheet.max_column + 1 
    for row in range(2, rows + 1):
        line = {}
        for col in range(1, columns):
            key = worksheet.cell(row=1, column=col).value
            value = worksheet.cell(row=row, column=col).value
            line[key] = value
        table.append(line)

    return table

def set_value(mydict:dict, paths:list, value) -> None:
    """set value in a dict

    Parameters
    ----------
    mydict : dict
        the 'dict' in which 'the value is set.
    paths : list
        path to use to set value on 'correct' position
    value : 
        the new value
    """    
    # write value to nested dict
    # we split the path by using '__'
    parts = paths.split('__')
    for part in parts[0:-1]:
        # add {} if item does not exists
        # this loop create an empty path
        mydict = mydict.setdefault(part, {})
    # at last write value to dict
    mydict[parts[-1]] = value

def get_prefix_path(prefixe:list, ip:str) -> list:
    """return prefix path of ip

    Parameters
    ----------
    prefixe : list
        the list of ALL prefixes
    ip : str
        the IP address

    Returns
    -------
    list
        list of prefixes that include the IP address
    """
    prefix_path = []
    pyt = pytricia.PyTricia()

    # build pytricia tree
    for prefix_ip in prefixe:
        pyt.insert(prefix_ip, prefix_ip)

    try:
        prefix = pyt.get(ip)
    except Exception:
        logger.info('prefix not found; using 0.0.0.0/0')
        prefix = "0.0.0.0/0"
    prefix_path.append(prefix)

    parent = pyt.parent(prefix)
    while (parent):
        prefix_path.append(parent)
        parent = pyt.parent(parent)
    return prefix_path[::-1]

def calculate_md5(row:list):
    """calculate MD5 value of all columns in a row

    Parameters
    ----------
    row : list
        the list that contains all columns (values)

    Returns
    -------
    md5 
        the MD5 hash
    """
    data = ""
    for cell in row:
        if isinstance(cell, list):
            my_list = ''.join(cell)
            data += my_list
        elif cell is None or cell == 'null':
            pass
        else:
            data += cell
    return hashlib.md5(data.encode('utf-8')).hexdigest()

def get_date(date:str) -> str:
    """return string of date

    Parameters
    ----------
    date : str
        date to convert like 'today', 'this_week'

    Returns
    -------
    str
        the converted date
    """    
    today = datetime.date.today()

    if date == 'today':
        now = datetime.datetime.now()
        return now.strftime("%Y-%m-%d 00:00:00")
    elif date == 'this_week':
        # our week begins on monday NOT sunday
        monday = today - datetime.timedelta(days=today.weekday())
        return monday.strftime("%Y-%m-%d 00:00:00")
    elif date == 'last_week' or date == 'last_seven_days':
        # that is last seven days
        last_week = today - datetime.timedelta(days=7)
        return last_week.strftime("%Y-%m-%d 00:00:00")
    elif date == 'this_month':
        day_num = today.strftime("%d")
        this_month = today - datetime.timedelta(days=int(day_num) - 1)
        return this_month.strftime("%Y-%m-%d 00:00:00")
    else:
        return None

def analyze_nornir_result(result) -> dict:
    """analyze nornir result and return dict

    Parameters
    ----------
    result : 
        nornir result

    Returns
    -------
    analysis : dict
        the analyzed result
    """
    logger.debug('analyzing results')
    analysis = {}
    hosts = result.keys()
    logger.debug(f'got result(s) from {",".join(hosts)}')
    for host in hosts:
        analysis[host] = {}
        nn_tasks = len(result[host])
        logger.bind(extra=host).debug(f'got {nn_tasks} results')
        for i in range(0, nn_tasks):
            name = result[host][i].name
            failed = result[host][i].failed
            changed = result[host][i].changed
            analysis[host][name] = {'failed': failed,
                                    'changed': changed}
            logger.bind(extra=host).debug(f'{i}: name: {name}')
            logger.bind(extra=host).debug(f'{i}: failed: {failed}')
            logger.bind(extra=host).debug(f'{i}: changed: {changed}')

    return analysis

def write_mail(email_message:str, server_properties:dict):
    """write email

    Parameters
    ----------
    email_message : str
        email message
    server_properties : dict
        server properties

    Returns
    -------
    success : bool
        True if success
    """
    smtp_server = server_properties.get('server')
    username = server_properties.get('username')
    password = server_properties.get('password')

    smtp = smtplib.SMTP(smtp_server, server_properties.get('port', 587))
    #smtp.set_debuglevel(1)
    smtp.ehlo()
    smtp.starttls()
    smtp.login(username, password)
    smtp.sendmail(email_message['From'] , email_message['To'], email_message.as_string())
    smtp.quit()

    return True

def flatten_dict_containing_lists(dictionary:dict, parent='') -> dict:
    """this function flattens a dict that contains lists"""
    for key, value in dictionary.items():
        if isinstance(value, dict):
            for k, v in flatten_dict_containing_lists(value, parent + key + '.'):
                yield  k, v
        elif isinstance(value, list):
            x = 0
            for i in value:
                for k, v in flatten_dict_containing_lists(i, key + f'[{x}].'):
                    x += 1
                    yield k, v
        else:
            yield parent + key, value

    def find_in_line(self, key:str, lookup:str, value:str, line:str) -> bool:
        """find value in line

        n - not equal to (negation)
        ic - case-insensitive contains (*)
        c - case-sensitive contains (*)
        ie - case-insensitive exact match (*)
        isw - case-insensitive starts-with
        iew - case-insensitive ends-with
        re - case-sensitive regular expression match
        nic - negated case-insensitive contains
        nisw - negated case-insensitive starts-with
        niew - negated case-insensitive ends-with
        nie - negated case-insensitive exact match
        nre - negated case-sensitive regular expression match
        ire - case-insensitive regular expression match
        nire - negated case-insensitive regular expression match

        Parameters
        ----------
        key : str
            what to do (currently only match is supported)
        lookup : str
            the lookup type (ie, ic, c, n, ...)
        value : str
            the value to look for
        line : str
            the line to search in

        Returns
        -------
        bool
            true if found, false otherwise
        """

        # logger.debug(f'key: {key} lookup: {lookup} value: {value} line: {line}')
        if key == 'match':
            if lookup == "ie":
                # case-insensitive exact match
                if line.lower() == value.lower():
                    return True
            elif lookup == "ic":
                # case-insensitive contains
                if value.lower() in line.lower():
                    return True
            elif lookup == "c":
            # case-sensitive contains
                if value in line:
                    return True
            elif lookup == "isw":
                # case-insensitive starts-with
                if line.lower().startswith(value.lower()):
                    return True
            elif lookup == "iew":
                # case-insensitive ends-with
                if line.lower().endswith(value.lower()):
                    return True
            elif lookup == "re":
                # case-sensitive regular expression match
                # compile regular expression
                pattern = re.compile(value)
                if pattern.search(line):
                    return True
            elif lookup == "nic":
                # negated case-insensitive contains
                if value.lower() not in line.lower():
                    return True
            elif lookup == "nisw":
                # negated case-insensitive starts-with
                if not line.lower().startswith(value.lower()):
                    return True
            elif lookup == "niew":
                # negated case-insensitive ends-with
                if not line.lower().endswith(value.lower()):
                    return True
            elif lookup == "nie":
                # negated case-insensitive exact match
                if line.lower() != value.lower():
                    return True
            elif lookup == "nre":
                # negated case-sensitive regular expression match
                # compile regular expression
                pattern = re.compile(value)
                if not pattern.search(line):
                    return True
            elif lookup == "ire":    
                # case-insensitive regular expression match
                # compile regular expression
                pattern = re.compile(value, re.IGNORECASE)
                if pattern.search(line):
                    return True
            elif lookup == "nire":
                # negated case-insensitive regular expression match
                # compile regular expression
                pattern = re.compile(value, re.IGNORECASE)
                if not pattern.search(line):
                    return True
            else:
                if line == value:
                    return True

        return False
